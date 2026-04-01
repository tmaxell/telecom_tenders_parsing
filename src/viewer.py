#!/usr/bin/env python3
"""
Tender Viewer & Exporter
========================
Просмотр, поиск, статистика и экспорт собранных тендеров.

Примеры:
    python -m src.viewer --last 20
    python -m src.viewer --stats
    python -m src.viewer --search "IMEI"
    python -m src.viewer --export csv
    python -m src.viewer --export excel
    python -m src.viewer --export html
    python -m src.viewer --matched --export csv
    python -m src.viewer --source bicotender --last 50
    python -m src.viewer --interactive
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import yaml

logger = logging.getLogger("viewer")


# ══════════════════════════════════════════════════════════════════
#  Database reader (read-only)
# ══════════════════════════════════════════════════════════════════
class TenderReader:
    """Read-only доступ к БД тендеров."""

    def __init__(self, db_path: str = "data/tenders.db"):
        if not Path(db_path).exists():
            print(f"❌ База данных не найдена: {db_path}")
            print("   Сначала запустите сбор: python -m src.collector")
            sys.exit(1)

        self.conn = sqlite3.connect(db_path)
        self.conn.row_factory = sqlite3.Row

    def query(self, sql: str, params: tuple = ()) -> list[dict]:
        rows = self.conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]

    def count(self, table: str = "raw_tenders") -> int:
        return self.conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]

    # ── Selections ────────────────────────────────────────────────
    def get_last(self, n: int = 20, source: str = None) -> list[dict]:
        where = "WHERE source_id = ?" if source else ""
        params = (source,) if source else ()
        sql = f"""
            SELECT ocid, source_id, country, title, buyer_name,
                   value_amount, value_currency, status,
                   procurement_method, date_published,
                   tender_period_end, collected_at
            FROM raw_tenders
            {where}
            ORDER BY collected_at DESC
            LIMIT ?
        """
        return self.query(sql, (*params, n))

    def search(self, text: str, limit: int = 50) -> list[dict]:
        pattern = f"%{text}%"
        sql = """
            SELECT ocid, source_id, country, title, description,
                   buyer_name, value_amount, value_currency,
                   date_published, tender_period_end, status
            FROM raw_tenders
            WHERE title LIKE ? OR description LIKE ?
                  OR items_text LIKE ? OR buyer_name LIKE ?
            ORDER BY date_published DESC
            LIMIT ?
        """
        return self.query(sql, (pattern, pattern, pattern, pattern, limit))

    def get_matched(self, product: str = None, limit: int = 500) -> list[dict]:
        if product:
            sql = """
                SELECT * FROM matched_tenders
                WHERE matched_product = ?
                ORDER BY date_published DESC LIMIT ?
            """
            return self.query(sql, (product, limit))
        else:
            sql = """
                SELECT * FROM matched_tenders
                ORDER BY date_published DESC LIMIT ?
            """
            return self.query(sql, (limit,))

    def get_all_raw(self, source: str = None) -> list[dict]:
        if source:
            sql = """
                SELECT ocid, source_id, country, title, description,
                       buyer_name, value_amount, value_currency,
                       status, procurement_method,
                       date_published, tender_period_end,
                       items_text, collected_at
                FROM raw_tenders WHERE source_id = ?
                ORDER BY date_published DESC
            """
            return self.query(sql, (source,))
        else:
            sql = """
                SELECT ocid, source_id, country, title, description,
                       buyer_name, value_amount, value_currency,
                       status, procurement_method,
                       date_published, tender_period_end,
                       items_text, collected_at
                FROM raw_tenders ORDER BY date_published DESC
            """
            return self.query(sql)

    def get_stats(self) -> dict:
        stats = {
            "total_raw": self.count("raw_tenders"),
            "total_matched": self.count("matched_tenders"),
        }

        stats["by_source"] = self.query("""
            SELECT source_id, country, COUNT(*) as cnt,
                   MIN(date_published) as earliest,
                   MAX(date_published) as latest
            FROM raw_tenders GROUP BY source_id
            ORDER BY cnt DESC
        """)

        stats["by_status"] = self.query("""
            SELECT status, COUNT(*) as cnt
            FROM raw_tenders GROUP BY status
            ORDER BY cnt DESC
        """)

        stats["by_country"] = self.query("""
            SELECT country, COUNT(*) as cnt
            FROM raw_tenders GROUP BY country
            ORDER BY cnt DESC
        """)

        stats["by_product"] = self.query("""
            SELECT matched_product, COUNT(*) as cnt
            FROM matched_tenders GROUP BY matched_product
            ORDER BY cnt DESC
        """)

        stats["top_buyers"] = self.query("""
            SELECT buyer_name, COUNT(*) as cnt
            FROM raw_tenders
            WHERE buyer_name != ''
            GROUP BY buyer_name
            ORDER BY cnt DESC LIMIT 15
        """)

        stats["collection_state"] = self.query("""
            SELECT * FROM collection_state ORDER BY last_collected DESC
        """)

        return stats

    def close(self):
        self.conn.close()


# ══════════════════════════════════════════════════════════════════
#  Formatters
# ══════════════════════════════════════════════════════════════════

# ── Console table ─────────────────────────────────────────────────
def print_table(rows: list[dict], columns: list[str] = None,
                max_col_width: int = 40):
    """Красивая таблица в терминале."""
    if not rows:
        print("  (нет данных)")
        return

    if not columns:
        columns = list(rows[0].keys())

    # exclude raw_json and very long fields from console
    skip = {"raw_json", "items_text", "full_text", "description"}
    columns = [c for c in columns if c not in skip]

    # Ширины колонок
    widths = {}
    for col in columns:
        values = [str(row.get(col, ""))[:max_col_width] for row in rows]
        widths[col] = max(len(col), max(len(v) for v in values))
        widths[col] = min(widths[col], max_col_width)

    # Header
    header = " │ ".join(col.ljust(widths[col])[:widths[col]] for col in columns)
    sep = "─┼─".join("─" * widths[col] for col in columns)
    print(f" {header}")
    print(f" {sep}")

    # Rows
    for row in rows:
        line = " │ ".join(
            str(row.get(col, ""))[:widths[col]].ljust(widths[col])
            for col in columns
        )
        print(f" {line}")


def print_tender_detail(row: dict):
    """Подробный вывод одного тендера."""
    print(f"\n{'━' * 70}")
    print(f"  OCID       : {row.get('ocid', 'N/A')}")
    print(f"  Title      : {row.get('title', 'N/A')}")
    print(f"  Source     : {row.get('source_id', '')} ({row.get('country', '')})")
    print(f"  Buyer      : {row.get('buyer_name', 'N/A')}")
    print(f"  Price      : {row.get('value_amount', 'N/A')} {row.get('value_currency', '')}")
    print(f"  Status     : {row.get('status', 'N/A')}")
    print(f"  Law/Method : {row.get('procurement_method', 'N/A')}")
    print(f"  Published  : {row.get('date_published', 'N/A')}")
    print(f"  Deadline   : {row.get('tender_period_end', 'N/A')}")
    print(f"  Collected  : {row.get('collected_at', 'N/A')}")

    if row.get("matched_product"):
        print(f"  Product    : {row['matched_product']}")
        print(f"  Keywords   : {row.get('matched_keywords', '')}")
        print(f"  Snippet    : {row.get('match_snippet', '')[:200]}")

    desc = row.get("description", "")
    if desc:
        print(f"  Description: {desc[:300]}{'…' if len(desc)>300 else ''}")

    print(f"{'━' * 70}")


# ── Export functions ──────────────────────────────────────────────
def export_csv(rows: list[dict], path: str):
    """Экспорт в CSV."""
    if not rows:
        print("❌ Нет данных для экспорта")
        return

    exclude = {"raw_json"}
    fields = [k for k in rows[0].keys() if k not in exclude]

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fields})

    print(f"✔ Экспортировано {len(rows)} записей → {path}")


def export_json(rows: list[dict], path: str):
    """Экспорт в JSON."""
    if not rows:
        print("❌ Нет данных для экспорта")
        return

    exclude = {"raw_json"}
    clean = [{k: v for k, v in row.items() if k not in exclude} for row in rows]

    with open(path, "w", encoding="utf-8") as f:
        json.dump(clean, f, ensure_ascii=False, indent=2, default=str)

    print(f"✔ Экспортировано {len(rows)} записей → {path}")


def export_excel(rows: list[dict], path: str):
    """Экспорт в Excel (.xlsx). Требует openpyxl."""
    if not rows:
        print("❌ Нет данных для экспорта")
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ Для Excel-экспорта установите: pip install openpyxl")
        print("   Пока экспортирую в CSV …")
        export_csv(rows, path.replace(".xlsx", ".csv"))
        return

    exclude = {"raw_json"}
    fields = [k for k in rows[0].keys() if k not in exclude]

    wb = Workbook()
    ws = wb.active
    ws.title = "Tenders"

    # ── Стили ─────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                              fill_type="solid")
    header_alignment = Alignment(horizontal="center", wrap_text=True)

    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2",
                           fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Человекочитаемые заголовки
    header_labels = {
        "ocid": "ID",
        "source_id": "Источник",
        "country": "Страна",
        "title": "Название тендера",
        "description": "Описание",
        "buyer_name": "Заказчик",
        "buyer_id": "ИНН",
        "value_amount": "Сумма",
        "value_currency": "Валюта",
        "status": "Статус",
        "procurement_method": "Закон/Метод",
        "date_published": "Дата публикации",
        "tender_period_start": "Начало приёма",
        "tender_period_end": "Дедлайн",
        "items_text": "Позиции",
        "collected_at": "Дата сбора",
        "matched_product": "Продукт",
        "matched_keywords": "Ключевые слова",
        "match_snippet": "Контекст совпадения",
    }

    # ── Заголовки ─────────────────────────────────────────────────
    for col_idx, field_name in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx,
                       value=header_labels.get(field_name, field_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Данные ────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, field_name in enumerate(fields, 1):
            value = row_data.get(field_name, "")

            # Обрезаем очень длинные строки
            if isinstance(value, str) and len(value) > 500:
                value = value[:500] + "…"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Чередование цвета строк
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # ── Автоширина колонок ────────────────────────────────────────
    for col_idx, field_name in enumerate(fields, 1):
        max_len = len(header_labels.get(field_name, field_name))
        for row_data in rows[:100]:  # сэмплируем первые 100 строк
            val = str(row_data.get(field_name, ""))
            max_len = max(max_len, min(len(val), 60))

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 65)

    # ── Закрепить заголовок ───────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Автофильтр ────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    print(f"✔ Экспортировано {len(rows)} записей → {path}")


def export_html(rows: list[dict], path: str):
    """Экспорт в HTML с сортировкой и поиском (standalone файл)."""
    if not rows:
        print("❌ Нет данных для экспорта")
        return

    exclude = {"raw_json", "items_text"}
    fields = [k for k in rows[0].keys() if k not in exclude]

    header_labels = {
        "ocid": "ID", "source_id": "Источник", "country": "Страна",
        "title": "Название", "description": "Описание",
        "buyer_name": "Заказчик", "value_amount": "Сумма",
        "value_currency": "Валюта", "status": "Статус",
        "procurement_method": "Закон", "date_published": "Опубликовано",
        "tender_period_end": "Дедлайн", "matched_product": "Продукт",
        "matched_keywords": "Ключевые слова", "match_snippet": "Контекст",
        "collected_at": "Собрано",
    }

    # HTML с встроенным CSS и JS для сортировки/фильтрации
    html_parts = ["""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>Tender Monitor — Export</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, 'Segoe UI', Roboto, sans-serif;
         background: #f5f5f5; color: #333; padding: 20px; }
  h1 { margin-bottom: 10px; color: #2F5496; }
  .meta { color: #888; margin-bottom: 15px; }
  .controls { margin-bottom: 15px; }
  #searchInput {
    padding: 8px 14px; width: 400px; font-size: 14px;
    border: 1px solid #ccc; border-radius: 6px;
  }
  table { border-collapse: collapse; width: 100%; background: #fff;
          box-shadow: 0 1px 4px rgba(0,0,0,0.1); border-radius: 8px;
          overflow: hidden; }
  th { background: #2F5496; color: #fff; padding: 10px 8px;
       text-align: left; font-size: 12px; cursor: pointer;
       user-select: none; white-space: nowrap; }
  th:hover { background: #3a68b5; }
  th .arrow { margin-left: 4px; font-size: 10px; }
  td { padding: 8px; font-size: 12px; border-bottom: 1px solid #eee;
       max-width: 300px; overflow: hidden; text-overflow: ellipsis;
       white-space: nowrap; }
  td:hover { white-space: normal; overflow: visible; }
  tr:nth-child(even) { background: #fafafa; }
  tr:hover { background: #e8f0fe; }
  .amount { text-align: right; font-family: monospace; }
  .highlight { background: #fff3cd; }
  .count { font-weight: bold; color: #2F5496; }
</style>
</head>
<body>
<h1>📋 Tender Monitor</h1>
<p class="meta">Экспорт: """ + datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC") + f""" | Всего записей: <span class="count">{len(rows)}</span></p>
<div class="controls">
  <input id="searchInput" type="text" placeholder="🔍 Поиск по таблице …" onkeyup="filterTable()">
</div>
<table id="tenderTable">
<thead><tr>"""]

    for i, f in enumerate(fields):
        label = header_labels.get(f, f)
        html_parts.append(
            f'<th onclick="sortTable({i})">{label}<span class="arrow">⇅</span></th>'
        )

    html_parts.append("</tr></thead><tbody>")

    for row in rows:
        html_parts.append("<tr>")
        for f in fields:
            val = str(row.get(f, "") or "")
            if len(val) > 200:
                val = val[:200] + "…"
            # escape HTML
            val = val.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            cls = ' class="amount"' if f in ("value_amount",) else ""
            html_parts.append(f"<td{cls}>{val}</td>")
        html_parts.append("</tr>")

    html_parts.append("""</tbody></table>
<script>
function filterTable() {
  const input = document.getElementById('searchInput').value.toLowerCase();
  const rows = document.querySelectorAll('#tenderTable tbody tr');
  rows.forEach(row => {
    const text = row.textContent.toLowerCase();
    row.style.display = text.includes(input) ? '' : 'none';
  });
}
function sortTable(colIdx) {
  const table = document.getElementById('tenderTable');
  const tbody = table.querySelector('tbody');
  const rows = Array.from(tbody.rows);
  const dir = table.dataset.sortDir === 'asc' ? 'desc' : 'asc';
  table.dataset.sortDir = dir;
  rows.sort((a, b) => {
    let va = a.cells[colIdx].textContent.trim();
    let vb = b.cells[colIdx].textContent.trim();
    const na = parseFloat(va.replace(/[^\\d.-]/g, ''));
    const nb = parseFloat(vb.replace(/[^\\d.-]/g, ''));
    if (!isNaN(na) && !isNaN(nb)) { va = na; vb = nb; }
    if (va < vb) return dir === 'asc' ? -1 : 1;
    if (va > vb) return dir === 'asc' ? 1 : -1;
    return 0;
  });
  rows.forEach(row => tbody.appendChild(row));
}
</script>
</body></html>""")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(html_parts))

    print(f"✔ Экспортировано {len(rows)} записей → {path}")
    print(f"  Откройте в браузере: open {path}")


# ══════════════════════════════════════════════════════════════════
#  Statistics printer
# ══════════════════════════════════════════════════════════════════
def print_stats(reader: TenderReader):
    stats = reader.get_stats()

    print(f"\n{'═' * 70}")
    print(f"  📊 СТАТИСТИКА TENDER MONITOR")
    print(f"{'═' * 70}")
    print(f"  Всего тендеров в БД      : {stats['total_raw']}")
    print(f"  Совпадений по keywords   : {stats['total_matched']}")

    print(f"\n  📁 По источникам:")
    print_table(stats["by_source"])

    print(f"\n  🌍 По странам:")
    print_table(stats["by_country"])

    print(f"\n  📌 По статусам:")
    print_table(stats["by_status"])

    if stats["by_product"]:
        print(f"\n  🎯 Совпадения по продуктам:")
        print_table(stats["by_product"])

    if stats["top_buyers"]:
        print(f"\n  🏢 Топ-15 заказчиков:")
        print_table(stats["top_buyers"])

    if stats["collection_state"]:
        print(f"\n  ⏱ Последний сбор:")
        print_table(stats["collection_state"])

    print(f"{'═' * 70}\n")


# ══════════════════════════════════════════════════════════════════
#  Interactive mode
# ══════════════════════════════════════════════════════════════════
def interactive_mode(reader: TenderReader):
    """Интерактивный режим просмотра."""
    print(f"\n{'═' * 70}")
    print("  🔍 ИНТЕРАКТИВНЫЙ ПРОСМОТР ТЕНДЕРОВ")
    print(f"{'═' * 70}")
    print("  Команды:")
    print("    last [N]          — последние N тендеров (по умолчанию 20)")
    print("    search <текст>    — поиск по тексту")
    print("    matched [product] — совпадения по ключевым словам")
    print("    stats             — статистика")
    print("    detail <ocid>     — подробно о тендере")
    print("    export <format>   — экспорт (csv/json/excel/html)")
    print("    quit              — выход")
    print(f"{'─' * 70}\n")

    while True:
        try:
            cmd = input("tender> ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nДо свидания!")
            break

        if not cmd:
            continue

        parts = cmd.split(maxsplit=1)
        action = parts[0].lower()
        arg = parts[1] if len(parts) > 1 else ""

        if action in ("quit", "exit", "q"):
            print("До свидания!")
            break

        elif action == "last":
            n = int(arg) if arg.isdigit() else 20
            rows = reader.get_last(n)
            print(f"\n  Последние {n} тендеров ({len(rows)} найдено):\n")
            print_table(rows)

        elif action == "search":
            if not arg:
                print("  Использование: search <текст>")
                continue
            rows = reader.search(arg)
            print(f"\n  Поиск '{arg}': {len(rows)} результатов\n")
            print_table(rows)

        elif action == "matched":
            rows = reader.get_matched(product=arg if arg else None)
            label = f"по продукту '{arg}'" if arg else "все"
            print(f"\n  Совпадения ({label}): {len(rows)}\n")
            print_table(rows)

        elif action == "stats":
            print_stats(reader)

        elif action == "detail":
            if not arg:
                print("  Использование: detail <ocid>")
                continue
            rows = reader.query(
                "SELECT * FROM raw_tenders WHERE ocid = ?", (arg,)
            )
            if rows:
                print_tender_detail(rows[0])
            else:
                # поиск по частичному совпадению
                rows = reader.query(
                    "SELECT * FROM raw_tenders WHERE ocid LIKE ? LIMIT 5",
                    (f"%{arg}%",)
                )
                if rows:
                    for r in rows:
                        print_tender_detail(r)
                else:
                    print(f"  Тендер '{arg}' не найден")

        elif action == "export":
            fmt = arg.lower() if arg else "csv"
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            out_dir = "data/exports"
            Path(out_dir).mkdir(parents=True, exist_ok=True)

            # спрашиваем что экспортировать
            print("  Что экспортировать?")
            print("    1 — все собранные тендеры")
            print("    2 — только совпадения по ключевым словам")
            choice = input("  Выбор (1/2): ").strip()

            if choice == "2":
                rows = reader.get_matched()
                prefix = "matched"
            else:
                rows = reader.get_all_raw()
                prefix = "all_tenders"

            path = f"{out_dir}/{prefix}_{ts}"
            if fmt == "csv":
                export_csv(rows, f"{path}.csv")
            elif fmt == "json":
                export_json(rows, f"{path}.json")
            elif fmt in ("excel", "xlsx"):
                export_excel(rows, f"{path}.xlsx")
            elif fmt == "html":
                export_html(rows, f"{path}.html")
            else:
                print(f"  Неизвестный формат: {fmt}")

        else:
            print(f"  Неизвестная команда: {action}")


# ══════════════════════════════════════════════════════════════════
#  Entry point
# ══════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="Tender Viewer & Exporter",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры:
  python -m src.viewer --last 20
  python -m src.viewer --stats
  python -m src.viewer --search "IMEI"
  python -m src.viewer --matched --product CEIR
  python -m src.viewer --export excel
  python -m src.viewer --export html --matched
  python -m src.viewer --interactive
        """,
    )

    parser.add_argument("--last", type=int, metavar="N",
                        help="Показать последние N тендеров")
    parser.add_argument("--search", type=str,
                        help="Поиск по тексту")
    parser.add_argument("--matched", action="store_true",
                        help="Показать/экспортировать только совпадения")
    parser.add_argument("--product", type=str,
                        help="Фильтр по продукту (CEIR, CVM, …)")
    parser.add_argument("--source", type=str,
                        help="Фильтр по источнику (bicotender, …)")
    parser.add_argument("--stats", action="store_true",
                        help="Показать статистику")
    parser.add_argument("--detail", type=str, metavar="OCID",
                        help="Подробно о конкретном тендере")
    parser.add_argument("--export", type=str,
                        choices=["csv", "json", "excel", "html", "all"],
                        help="Экспорт в файл")
    parser.add_argument("--interactive", "-i", action="store_true",
                        help="Интерактивный режим")
    parser.add_argument("--db", type=str, default="data/tenders.db",
                        help="Путь к БД")

    args = parser.parse_args()

    # Если нет аргументов — показать справку
    if len(sys.argv) == 1:
        parser.print_help()
        print("\n  💡 Быстрый старт: python -m src.viewer --stats")
        print("  💡 Интерактив   : python -m src.viewer -i\n")
        return

    reader = TenderReader(args.db)

    try:
        if args.interactive:
            interactive_mode(reader)
            return

        if args.stats:
            print_stats(reader)

        if args.last:
            rows = reader.get_last(args.last, source=args.source)
            print(f"\n  Последние {args.last} тендеров:\n")
            print_table(rows)

        if args.search:
            rows = reader.search(args.search)
            print(f"\n  Поиск '{args.search}': {len(rows)} результатов\n")
            print_table(rows)

        if args.detail:
            rows = reader.query(
                "SELECT * FROM raw_tenders WHERE ocid LIKE ?",
                (f"%{args.detail}%",),
            )
            for r in rows:
                print_tender_detail(r)

        if args.matched and not args.export:
            rows = reader.get_matched(product=args.product)
            print(f"\n  Совпадения: {len(rows)}\n")
            print_table(rows)

        if args.export:
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            out_dir = "data/exports"
            Path(out_dir).mkdir(parents=True, exist_ok=True)

            if args.matched:
                rows = reader.get_matched(product=args.product)
                prefix = "matched"
            else:
                rows = reader.get_all_raw(source=args.source)
                prefix = "all_tenders"

            formats = ["csv", "json", "excel", "html"] if args.export == "all" else [args.export]

            for fmt in formats:
                path = f"{out_dir}/{prefix}_{ts}"
                if fmt == "csv":
                    export_csv(rows, f"{path}.csv")
                elif fmt == "json":
                    export_json(rows, f"{path}.json")
                elif fmt == "excel":
                    export_excel(rows, f"{path}.xlsx")
                elif fmt == "html":
                    export_html(rows, f"{path}.html")

    finally:
        reader.close()


if __name__ == "__main__":
    main()